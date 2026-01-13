#!/usr/bin/env python3
"""
Extract open tab URLs and workspace favorites from Microsoft Edge Workspace .edge files.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
import zlib
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


GZIP_MAGIC = b"\x1f\x8b"
INTERNAL_SCHEMES = {"about", "chrome", "edge", "file", "microsoft-edge"}


def default_input_path() -> str:
    if getattr(sys, "frozen", False):
        return str(Path(sys.executable).resolve().parent)
    return "."


def iter_gzip_offsets(data: bytes) -> Iterable[int]:
    start = 0
    while True:
        idx = data.find(GZIP_MAGIC, start)
        if idx == -1:
            return
        yield idx
        start = idx + 1


def decompress_payloads(data: bytes) -> list[bytes]:
    payloads: list[bytes] = []
    seen: set[bytes] = set()
    for idx in iter_gzip_offsets(data):
        try:
            decompressor = zlib.decompressobj(16 + zlib.MAX_WBITS)
            out = decompressor.decompress(data[idx:])
            if not decompressor.eof or not out:
                continue
            digest = hashlib.sha256(out).digest()
            if digest in seen:
                continue
            seen.add(digest)
            payloads.append(out)
        except zlib.error:
            continue
    return payloads


def iter_json_objects(text: str) -> Iterable[Any]:
    decoder = json.JSONDecoder()
    idx = 0
    while idx < len(text):
        if text[idx] not in "{[":
            idx += 1
            continue
        try:
            obj, end = decoder.raw_decode(text, idx)
        except json.JSONDecodeError:
            idx += 1
            continue
        yield obj
        idx = end


def iter_content_objects(obj: Any) -> Iterable[dict[str, Any]]:
    if isinstance(obj, dict):
        content = obj.get("content")
        if isinstance(content, dict):
            yield content
        for value in obj.values():
            yield from iter_content_objects(value)
    elif isinstance(obj, list):
        for item in obj:
            yield from iter_content_objects(item)
    elif isinstance(obj, str):
        candidate = obj.strip()
        if candidate.startswith("{") or candidate.startswith("["):
            try:
                nested = json.loads(candidate)
            except Exception:
                return
            yield from iter_content_objects(nested)


def typed_value(value: Any) -> Any:
    if isinstance(value, dict) and "value" in value:
        return value.get("value")
    return value


def extract_tabs_from_content(content: dict[str, Any]) -> list[dict[str, str]]:
    links: list[dict[str, str]] = []
    webcontents = (
        content.get("subdirectories", {})
        .get("tabstripmodel", {})
        .get("subdirectories", {})
        .get("webcontents", {})
        .get("subdirectories", {})
    )
    if not isinstance(webcontents, dict):
        return links

    for tab_data in webcontents.values():
        if not isinstance(tab_data, dict):
            continue
        storage = tab_data.get("storage", {})
        current_index = typed_value(storage.get("currentNavigationIndex"))
        if current_index is None:
            continue
        nav_stack = (
            tab_data.get("subdirectories", {})
            .get("navigationStack", {})
            .get("subdirectories", {})
        )
        if not isinstance(nav_stack, dict) or not nav_stack:
            continue
        current_key = str(current_index)
        entry = nav_stack.get(current_key)
        if not entry:
            numeric_keys = [int(key) for key in nav_stack.keys() if str(key).isdigit()]
            if numeric_keys:
                entry = nav_stack.get(str(max(numeric_keys)))
        if not entry:
            continue
        entry_storage = entry.get("storage", {})
        url = ""
        for key in ("virtualUrl", "originalRequestUrl", "url"):
            value = typed_value(entry_storage.get(key))
            if isinstance(value, str) and value:
                url = value
                break
        if not url:
            continue
        title_value = typed_value(entry_storage.get("title"))
        title = title_value if isinstance(title_value, str) else ""
        links.append({"url": url, "title": title})

    return links


def extract_favorites_from_content(content: dict[str, Any]) -> list[dict[str, str]]:
    links: list[dict[str, str]] = []
    favorites = content.get("subdirectories", {}).get("favorites", {})
    if not isinstance(favorites, dict):
        return links
    storage = favorites.get("storage", {})
    if not isinstance(storage, dict):
        return links

    for entry in storage.values():
        node = typed_value(entry)
        if not isinstance(node, dict):
            continue
        node_type = node.get("nodeType")
        url = node.get("url")
        if str(node_type) != "1" or not isinstance(url, str) or not url:
            continue
        title = node.get("title")
        links.append({"url": url, "title": title if isinstance(title, str) else ""})

    return links


def extract_workspace_data(payloads: Iterable[bytes]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    text = b"\n".join(payloads).decode("utf-8", errors="ignore")
    clean = "".join(ch if ord(ch) >= 0x20 else " " for ch in text)

    tabs: list[dict[str, str]] = []
    favorites: list[dict[str, str]] = []

    for obj in iter_json_objects(clean):
        for content in iter_content_objects(obj):
            tabs.extend(extract_tabs_from_content(content))
            favorites.extend(extract_favorites_from_content(content))

    return tabs, favorites


def iter_edge_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if input_path.is_dir():
        return sorted(input_path.glob("*.edge"))
    raise FileNotFoundError(f"Input path not found: {input_path}")


def filter_links(
    links: list[dict[str, str]], exclude_schemes: set[str]
) -> list[dict[str, str]]:
    if not exclude_schemes:
        return links
    filtered: list[dict[str, str]] = []
    for link in links:
        url = link.get("url", "")
        scheme = url.split(":", 1)[0].lower() if ":" in url else ""
        if scheme in exclude_schemes:
            continue
        filtered.append(link)
    return filtered


def resolve_output_path(input_path: Path, output: str | None) -> str:
    if output:
        return output
    base_dir = input_path if input_path.is_dir() else input_path.parent
    return str(base_dir / "edge_workspace_links.xlsx")


def write_output(
    rows: list[dict[str, str]],
    summary_rows: list[tuple[str, int]],
    file_rows: list[dict[str, str | int]],
    output_path: str,
) -> None:
    workbook = Workbook()

    links_sheet = workbook.active
    links_sheet.title = "Links"
    links_sheet.append(["workspace_file", "source", "url", "title"])
    for row in rows:
        links_sheet.append([row["workspace_file"], row["source"], row["url"], row["title"]])
        url_cell = links_sheet.cell(row=links_sheet.max_row, column=3)
        url_value = url_cell.value
        if isinstance(url_value, str) and url_value:
            url_cell.hyperlink = url_value
            url_cell.style = "Hyperlink"

    summary_sheet = workbook.create_sheet("Summary Report")
    summary_sheet.append(["metric", "value"])
    for metric, value in summary_rows:
        summary_sheet.append([metric, value])

    per_file_sheet = workbook.create_sheet("Per File Report")
    per_file_sheet.append(["workspace_file", "open_tab_count", "favorite_count", "links_written"])
    for row in file_rows:
        per_file_sheet.append(
            [
                row["workspace_file"],
                row["open_tab_count"],
                row["favorite_count"],
                row["links_written"],
            ]
        )

    for sheet in (links_sheet, summary_sheet, per_file_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max_len + 2, 80)

    workbook.save(output_path)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract open tab URLs and workspace favorites from Edge Workspace .edge files."
    )
    parser.add_argument(
        "-i",
        "--input",
        default=default_input_path(),
        help=(
            "Path to a .edge file or a directory containing .edge files. "
            "Defaults to the script/exe location."
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .xlsx file path.",
    )
    parser.add_argument(
        "--exclude-schemes",
        nargs="*",
        default=[],
        help="URL schemes to exclude (example: edge chrome file).",
    )
    parser.add_argument(
        "--exclude-internal",
        action="store_true",
        help="Exclude internal browser URLs (edge, chrome, about, file).",
    )
    parser.add_argument(
        "--mode",
        choices=["both", "tabs", "favorites"],
        default="both",
        help="What to export to the Links sheet (default: both).",
    )
    parser.add_argument(
        "--sort",
        action="store_true",
        help="Sort output rows by workspace file and URL.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    if Workbook is None:
        print(
            "Missing dependency 'openpyxl'. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        return 2

    args = parse_args(argv)
    input_path = Path(args.input).expanduser()

    try:
        edge_files = iter_edge_files(input_path)
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    if not edge_files:
        print("No .edge files found in the input path.", file=sys.stderr)
        return 1

    exclude_schemes = {s.lower() for s in args.exclude_schemes}
    if args.exclude_internal:
        exclude_schemes.update(INTERNAL_SCHEMES)

    rows: list[dict[str, str]] = []
    per_file_rows: list[dict[str, str | int]] = []
    for path in edge_files:
        data = path.read_bytes()
        payloads = decompress_payloads(data)
        if not payloads:
            per_file_rows.append(
                {
                    "workspace_file": path.name,
                    "open_tab_count": 0,
                    "favorite_count": 0,
                    "links_written": 0,
                }
            )
            continue
        tabs, favorites = extract_workspace_data(payloads)
        tabs = filter_links(tabs, exclude_schemes)
        favorites = filter_links(favorites, exclude_schemes)

        def unique_by_url(links: list[dict[str, str]]) -> dict[str, str]:
            url_to_title: dict[str, str] = {}
            for link in links:
                url = link.get("url")
                if not isinstance(url, str) or not url:
                    continue
                title = link.get("title", "")
                if not isinstance(title, str):
                    title = ""
                if url not in url_to_title or (not url_to_title[url] and title):
                    url_to_title[url] = title
            return url_to_title

        tab_urls = unique_by_url(tabs) if args.mode in {"both", "tabs"} else {}
        favorite_urls = unique_by_url(favorites) if args.mode in {"both", "favorites"} else {}

        combined_urls: dict[str, tuple[str, str]] = {}
        for url, title in favorite_urls.items():
            combined_urls[url] = ("favorite", title)
        for url, title in tab_urls.items():
            if url in combined_urls:
                continue
            combined_urls[url] = ("tab", title)

        per_file_rows.append(
            {
                "workspace_file": path.name,
                "open_tab_count": len(tab_urls),
                "favorite_count": len(favorite_urls),
                "links_written": len(combined_urls),
            }
        )

        for url, (source, title) in combined_urls.items():
            rows.append(
                {
                    "workspace_file": path.name,
                    "source": source,
                    "url": url,
                    "title": title,
                }
            )

    if args.sort:
        rows.sort(
            key=lambda row: (
                row["workspace_file"],
                0 if row["source"] == "favorite" else 1,
                row["url"],
                row["title"],
            )
        )

    if args.sort:
        per_file_rows.sort(key=lambda row: row["workspace_file"])

    total_files = len(edge_files)
    files_with_links = sum(1 for row in per_file_rows if row["links_written"] > 0)
    files_with_tabs = sum(1 for row in per_file_rows if row["open_tab_count"] > 0)
    files_with_favorites = sum(1 for row in per_file_rows if row["favorite_count"] > 0)
    tabs_total = sum(int(row["open_tab_count"]) for row in per_file_rows)
    favorites_total = sum(int(row["favorite_count"]) for row in per_file_rows)
    links_total = len(rows)
    unique_urls = len({row["url"] for row in rows})
    summary_rows = [
        ("files_found", total_files),
        ("files_with_any_links", files_with_links),
        ("files_with_tabs", files_with_tabs),
        ("files_with_favorites", files_with_favorites),
        ("tabs_total", tabs_total),
        ("favorites_total", favorites_total),
        ("links_total", links_total),
        ("unique_urls", unique_urls),
    ]

    output_path = resolve_output_path(input_path, args.output)
    write_output(rows, summary_rows, per_file_rows, output_path)

    print(
        f"Wrote {len(rows)} links from {len(edge_files)} workspace file(s) to {output_path}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
